import io
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def build_excel(df, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Оплаты"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E4057")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    alt_fill = PatternFill("solid", fgColor="F2F5F8")
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin
            cell.alignment = left
            if fill:
                cell.fill = fill
    for col_idx, header in enumerate(headers, start=1):
        col_data = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(col_data.map(len).max(), len(header)) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def fmt_money(value):
    return f"{value:,.0f} ₽".replace(",", " ")

def pluralize_payments(n):
    if 11 <= n % 100 <= 19:
        return "оплат"
    r = n % 10
    if r == 1:
        return "оплата"
    if 2 <= r <= 4:
        return "оплаты"
    return "оплат"

def build_analytics_text(df, date_str):
    total_count = len(df)
    has_profit = "Заработано" in df.columns
    has_cost = "Оплачено" in df.columns
    total_profit = df["Заработано"].sum() if has_profit else 0
    total_cost = df["Оплачено"].sum() if has_cost else 0
    lines = [
        f"📊 <b>Аналитика за {date_str}:</b>",
        f"Всего оплат: <b>{total_count}</b>",
    ]
    if has_profit:
        lines.append(f"Сумма (чистая): <b>{fmt_money(total_profit)}</b>")
    if has_cost:
        lines.append(f"Сумма (с комиссией): <b>{fmt_money(total_cost)}</b>")
    if has_profit and total_count > 0:
        lines.append(f"Средний чек: <b>{fmt_money(total_profit / total_count)}</b>")
    if "user_utm_source" in df.columns:
        lines.append("")
        lines.append("🎯 <b>По источникам трафика:</b>")
        group_col = "Заработано" if has_profit else ("Оплачено" if has_cost else None)
        if group_col:
            grouped = df.groupby("user_utm_source", dropna=False).agg(count=("user_utm_source","count"), total=(group_col,"sum")).sort_values("total", ascending=False)
            for source, row in grouped.iterrows():
                count = int(row["count"])
                lines.append(f"— {source}: {count} {pluralize_payments(count)} / {fmt_money(row['total'])}")
        else:
            for source, count in df["user_utm_source"].value_counts().items():
                lines.append(f"— {source}: {count} {pluralize_payments(count)}")
    return "\n".join(lines)
